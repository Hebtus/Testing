# To give us access to the enter key, escape key...etc. ex: when I write something in the search bar and want to press enter:
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
import re
import pyperclip
from datetime import datetime
import csv
import openpyxl

import sys
sys.path.append(".") # To access modules in sibling directories

from Common_Files.Utilities import *
from Common_Files.RealReferences_Eventbrite import *

def event_list(driver, mode=0):
    # ---------------------------------------------------- To Event List --------------------------------------------------- #
    NavigateToEvents = find_my_element(driver,"XPATH",NAVIGATE_TO_EVENTS)
    NavigateToEvents.click()
    time.sleep(1)
    # ---------------------------------------------------- Testing Filters --------------------------------------------------- #
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON) 
    # Test : Choose filter "All events", count each type and then use filter for every type to check if count matches.
    # Also extract all information to compare with CSV export
    # Lists to extract information
    EventNames=[]
    EventDates=[]
    EventStatuses=[]
    EventTicketsSold=[]
    EventTicketsAvailable=[]
    FilterButton.click()
    time.sleep(1)
    AllEventsButton = find_my_element(driver,"XPATH", FILTER_ALL_EVENTS)
    AllEventsButton.click()
    time.sleep(2)
    AllEvents=find_my_elements(driver,"XPATH",EVENTS_LIST_ALL_OR_UPCOMING)
    AllEventsCount = len(AllEvents)
    DraftCount = 0
    OnsaleCount = 0
    EndedCount = 0
    for i in range(1,(AllEventsCount+1)):
        # List for event names
        NameStrXPATH = EVENT_TITLE_PART1 + str(i) + EVENT_TITLE_PART2
        NameTemp = find_my_element(driver,"XPATH",NameStrXPATH)
        NameText=NameTemp.get_attribute('innerHTML')
        EventNames.append(NameText)
        # List for event dates
        DatesXPATH = EVENT_DATE_PART1 + str(i) + EVENT_DATE_PART2
        DateTemp = find_my_element(driver,"XPATH",DatesXPATH)
        DateText=DateTemp.get_attribute('innerHTML')
        EventDates.append(DateText)
        # List for event statuses
        StatusStrXPATH = EVENT_DRAFT_ONSALE_ENDED_STATUS_PART1 + str(i) + EVENT_DRAFT_ONSALE_ENDED_STATUS_PART2
        StatusTemp = find_my_element(driver,"XPATH",StatusStrXPATH)
        StatusHTML = StatusTemp.get_attribute('innerHTML')
        # List for Tickets sold
        TicketsSoldXPATH=EVENT_SOLD_PART1 + str(i) + EVENT_SOLD_PART2
        SoldTemp = find_my_element(driver,"XPATH",TicketsSoldXPATH)
        TicketsText=SoldTemp.get_attribute('innerHTML') # ex: "5 / 20"
        TicketsTemp=TicketsText.split()
        EventTicketsSold.append(int(TicketsTemp[0]))
        # List for Tickets available
        AvailableNum=int(TicketsTemp[2]) - int(TicketsTemp[0])
        EventTicketsAvailable.append(AvailableNum)

        if "Draft" in StatusHTML:
            DraftCount+=1
            EventStatuses.append("Draft")
        elif "On Sale" in StatusHTML:
            OnsaleCount+=1
            EventStatuses.append("Live")
        elif "Event ended" in StatusHTML:
            EndedCount+=1
            EventStatuses.append("Past")

    # ----- UPCOMING EVENTS: ----- #
    # Check if number of events in the filter of "upcoming events" is of the same count
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON_GENERAL)
    FilterButton.click()
    time.sleep(1)
    UpcomingEventsButton = find_my_element(driver,"XPATH", FILTER_UPCOMING_EVENTS)
    UpcomingEventsButton.click()
    time.sleep(2)
    AllUpcomingEvents = find_my_elements(driver,"XPATH",EVENTS_LIST_ALL_OR_UPCOMING)
    if len(AllUpcomingEvents) != OnsaleCount:
        print("Error: Not all upcoming events are shown using the filter of upcoming events")

    # Check if dates of upcoming events are all upcoming dates
    for i in range(1,OnsaleCount+1):
        if(OnsaleCount == 1):
            EventDate = find_my_element(driver,"XPATH",EVENT_DATE_ONE)
        else:
            EVENT_DATE= EVENT_DATE_PART1 + str(i) + EVENT_DATE_PART2
            EventDate = find_my_element(driver,"XPATH",EVENT_DATE)
        FullEventDate = EventDate.get_attribute('innerHTML') # ex: April 9, 2023 at 7:00 PM EET
        FullEventDateSeperated= FullEventDate.split() # default is split at whitespaces
        FullEventDateSeperated[1]=(FullEventDateSeperated[1].split(","))[0] # to get rid of comma next to day
        MonthNum = datetime.strptime(FullEventDateSeperated[0], '%B').month # to cast month from string to int (Note: April -> 4 (not 04))
        # to change formats: 9/4/2023 7:00PM -> 09/04/2023 07:00PM
        # Day
        if int(FullEventDateSeperated[1])<10:
            DayStr='0'+FullEventDateSeperated[1]
        else:
            DayStr=FullEventDateSeperated[1]
        # Month
        if MonthNum<10:
            MonthStr='0'+str(MonthNum)
        else:
            MonthStr=str(MonthNum)
        # Time
        TimeStr=FullEventDateSeperated[4].split(':')
        HourStr=TimeStr[0]
        if int(HourStr)<10:
            HourStr='0'+HourStr

        # Notice: month/day/year not day/month/year
        DateNewFormat = MonthStr + '/' + DayStr + '/' + FullEventDateSeperated[2] + ' ' + HourStr + ':' + TimeStr[1] + FullEventDateSeperated[5] #ex: 08/11/2019 05:45PM
        #print(DateNewFormat)

        datetime_object = datetime.today()
        TodayNewFormat = datetime.strftime(datetime_object, '%m/%d/%Y %I:%M%p') # to put in format of 08/11/2019 05:45PM
        TodayDate = datetime.strptime(TodayNewFormat, '%m/%d/%Y %I:%M%p') # to cast from str to timeobject
        #print(TodayDate)

        EventDate = datetime.strptime(DateNewFormat, '%m/%d/%Y %I:%M%p') # to cast from str to timeobject
        #print(EventDate)

        if(TodayDate>EventDate):
            print("Error: An event with a past date is displayed in upcoming events")

    # ----- DRAFTS EVENTS: ----- #
    # Check if number of events in the filter of "drafts" is of the same count
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON)
    FilterButton.click()
    time.sleep(1)
    DraftEventsButton = find_my_element(driver,"XPATH", FILTER_DRAFTS_EVENTS)
    DraftEventsButton.click()
    time.sleep(2)
    AllDraftEvents = find_my_elements(driver,"XPATH",EVENTS_LIST_PAST_OR_DRAFT)
    if len(AllDraftEvents) != DraftCount:
        print("Error: Not all draft events are shown using the filter of drafts")
    
    # ----- PAST EVENTS: ----- #
    # Check if number of events in the filter of "past events" is of the same count
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON_GENERAL)
    FilterButton.click()
    time.sleep(1)
    PastEventsButton = find_my_element(driver,"XPATH", FILTER_PAST_EVENTS)
    PastEventsButton.click()
    time.sleep(2)
    AllPastEvents = find_my_elements(driver,"XPATH",EVENTS_LIST_PAST_OR_DRAFT)
    if len(AllPastEvents) != EndedCount:
        print("Error: Not all past events are shown using the filter of past events")
    # Check if dates of past events are all past dates
    for i in range(1,EndedCount+1):
        if(EndedCount == 1):
            EventDate = find_my_element(driver,"XPATH",EVENT_DATE_ONE)
        else:
            EVENT_DATE= EVENT_DATE_PART1 + str(i) + EVENT_DATE_PART2
            EventDate = find_my_element(driver,"XPATH",EVENT_DATE)
        FullEventDate = EventDate.get_attribute('innerHTML') # ex: Monday, March 20, 2023 at 9:30 PM EET
        FullEventDateSeperated= FullEventDate.split() # default is split at whitespaces
        FullEventDateSeperated[2]=(FullEventDateSeperated[2].split(","))[0] # to get rid of comma next to day
        MonthNum = datetime.strptime(FullEventDateSeperated[1], '%B').month # to cast month from string to int (Note: April -> 4 (not 04))
        # to change formats: 9/4/2023 7:00PM -> 09/04/2023 07:00PM
        # Day
        if int(FullEventDateSeperated[2])<10:
            DayStr='0'+FullEventDateSeperated[2]
        else:
            DayStr=FullEventDateSeperated[2]
        # Month
        if MonthNum<10:
            MonthStr='0'+str(MonthNum)
        else:
            MonthStr=str(MonthNum)
        # Time
        TimeStr=FullEventDateSeperated[5].split(':')
        HourStr=TimeStr[0]
        if int(HourStr)<10:
            HourStr='0'+HourStr

        # Notice: month/day/year not day/month/year
        DateNewFormat = MonthStr + '/' + DayStr + '/' + FullEventDateSeperated[3] + ' ' + HourStr + ':' + TimeStr[1] + FullEventDateSeperated[6] #ex: 08/11/2019 05:45PM

        datetime_object = datetime.today()
        TodayNewFormat = datetime.strftime(datetime_object, '%m/%d/%Y %I:%M%p') # to put in format of 08/11/2019 05:45PM
        TodayDate = datetime.strptime(TodayNewFormat, '%m/%d/%Y %I:%M%p') # to cast from str to timeobject

        EventDate = datetime.strptime(DateNewFormat, '%m/%d/%Y %I:%M%p') # to cast from str to timeobject
        #print(EventDate)

        if(TodayDate<EventDate):
            print("Error: An event with an upcoming date is displayed in past events")
    
    # ---------------------------------------------------- Clicking on event ---------------------------------------------------- #
    # First go to upcoming events
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON_GENERAL)
    FilterButton.click()
    time.sleep(1)
    UpcomingEventsButton = find_my_element(driver,"XPATH", FILTER_UPCOMING_EVENTS)
    UpcomingEventsButton.click()
    time.sleep(1)
    # Clicking on an event should bring you to its dashboard
    AllUpcomingEvents = find_my_elements(driver,"XPATH",EVENTS_LIST_ALL_OR_UPCOMING)
    AllUpcomingEvents[0].click()
    time.sleep(3)
    DashboardElement = driver.find_element(By.XPATH,EVENT_DASHBOARD_TITLE)
    time.sleep(1)
    if DashboardElement.is_displayed() == False:
        print("Error: clicking on an event does not take you to its dashboard")
    DashboardLinkField=find_my_element(driver,"XPATH",EVENT_DASHBOARD_LINK)
    DashboardLink = str(DashboardLinkField.text)
    driver.back()
    

    # ------------------------------------------------- Three dots next to event ------------------------------------------------ #
    # Test 1: Promote on Eventbrite
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(10)
    PromoteChoice=find_my_element(driver,"XPATH",EVENT_PROMOTE_ON_EVENTBRITE_CHOICE)
    PromoteChoice.click()
    time.sleep(3)
    # Check if there was page navigation to promote on Eventbrite page
    PromoteElement = driver.find_element(By.XPATH,EVENT_PROMOTE_ON_EVENTBRITE_TITLE)
    if PromoteElement.is_displayed == False:
        print("Error: Did not navigate to promote on eventbrite page upon choosing from three dots")
    time.sleep(1)
    # Navigate back
    driver.back()

    # Test 2: View Event
    time.sleep(1)
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(1)
    ViewChoice=find_my_element(driver,"XPATH",EVENT_VIEW_CHOICE)
    ViewChoice.click()
    time.sleep(3)
    # Check if there was page navigation to view event page
    ViewElement = find_my_element(driver,"XPATH",EVENT_VIEW_TITLE)
    if ViewElement == None:
        print("Error: Did not navigate to view event page upon choosing from three dots")
    else:
        time.sleep(1)
        # Navigate back
        driver.back()

    # Test 3: Edit Event
    time.sleep(1)
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(1)
    EditChoice=find_my_element(driver,"XPATH",EVENT_EDIT_CHOICE)
    EditChoice.click()
    time.sleep(3)
    # Check if there was page navigation to edit event page
    EditElement = find_my_element(driver,"XPATH",EVENT_EDIT_TITLE)
    if EditElement == None:
        print("Error: Did not navigate to edit event page upon choosing from three dots")
    else:
        time.sleep(1)
        # Navigate back
        driver.back()

    # Test 4: Copy URL and check that it is the same as the one in the dashboard (Eventbrite website did not handle this case)
    time.sleep(1)
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(1)
    CopyURLChoice=find_my_element(driver,"XPATH",EVENT_COPY_URL_CHOICE)
    CopyURLChoice.click()
    URL = pyperclip.paste()
    # Check if copied URL is a valid URL:
    if "https://www.eventbrite.com" not in  URL:
        print("Error: clicking on Copy URL does not provide a valid URL")
    # Compare copied URL with the one we got for the same event but from the dashboard (uncomment when in Eventbrite mock-up)
    # print(URL)
    # print(DashboardLink)
    # if URL != DashboardLink:
      #  print("Error: copied URL does not match the URL provided in the event's dashboard")

    # Test 5: Copy Event
    time.sleep(1)
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(1)
    CopyChoice=find_my_element(driver,"XPATH",EVENT_COPY_EVENT_CHOICE)
    CopyChoice.click()
    time.sleep(3)
    # Check if there was page navigation to copy event page
    CopyElement = find_my_element(driver,"XPATH",EVENT_COPY_TITLE)
    if CopyElement == None:
        print("Error: Did not navigate to copy event page upon choosing from three dots")
    else:
        time.sleep(1)
        # Navigate back
        driver.back()

    # Test 6: Delete event
    # Go to draft events
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON_GENERAL)
    FilterButton.click()
    time.sleep(1)
    DraftEventsButton = find_my_element(driver,"XPATH", FILTER_DRAFTS_EVENTS)
    DraftEventsButton.click()
    time.sleep(2)
    if(DraftCount == 1):
        EventDeleted=find_my_element(driver,"XPATH",EVENT_DELETED_TITLE_ONE)
    else:
        EventDeleted=find_my_element(driver,"XPATH",EVENT_DELETED_TITLE)     
    ThreeDots=find_my_element(driver,"ID",THREE_DOTS_EVENT_BUTTON)
    ThreeDots.click()
    time.sleep(1)
    DeleteChoice=find_my_element(driver,"XPATH",EVENT_DELETE_CHOICE)
    DeleteChoice.click()
    time.sleep(3)
    alert = driver.switch_to.alert
    alert.accept()

    # Remove event from lists
    DeleteIndex = EventNames.index(EventDeleted.get_attribute('innerHTML'))
    del EventNames[DeleteIndex]
    del EventDates[DeleteIndex]
    del EventStatuses[DeleteIndex]
    del EventTicketsSold[DeleteIndex]
    del EventTicketsAvailable[DeleteIndex]

    time.sleep(5) # since page will refresh automatically

    # after refreshing it will be on page with filter "upcoming events" so choose all events filter and check that the count was decremented
    FilterButton = find_my_element(driver,"XPATH",FILTER_EVENTS_BUTTON)
    FilterButton.click()
    AllEventsButton = find_my_element(driver,"XPATH", FILTER_ALL_EVENTS)
    AllEventsButton.click()
    time.sleep(2)
    AllEvents=find_my_elements(driver,"XPATH",EVENTS_LIST_ALL_OR_UPCOMING)
    if(len(AllEvents) != (AllEventsCount-1)):
        print("Error: Deleted event is still displayed")
    # update draft and all events counts
    DraftCount -= 1
    AllEventsCount-=1

    # ------------------------------------------------- Testing CSV Export ------------------------------------------------ #
    CSVExport = find_my_element(driver,"XPATH",CSV_EXPORT_BUTTON)
    driver.execute_script("arguments[0].scrollIntoView();",CSVExport)
    CSVExport.click()
    time.sleep(10)
    # Change file extension to .xlsx
    csv_to_excel("C:/Users/MALAK/Downloads/MyEvents.csv","C:/Users/MALAK/Downloads/MyEvents.xlsx")
    Path="C:/Users/MALAK/Downloads/MyEvents.xlsx"
    WorkBook=openpyxl.load_workbook(Path)
    Sheet = WorkBook.active
    Rows=Sheet.max_row
    Cols=Sheet.max_column
    # Open file with the extension .xlsx and compare its values with the information previously retrieved for all events
    for r in range(1,Rows+1):
        for c in range(1,Cols+1):
            Value = Sheet.cell(row=r,column=c).value
            #print(Value,end="    ") # un-comment to print table
            if(r != 1):
                if(c == 1):
                   K = EventNames.index(Value) # Since order of rows in excel is different from in all events filter
                if(c == 2):
                    # format from May 3, 2023 at 7:00 PM EET -> May 3, 2023 7:00 PM
                    EventDateTemp = EventDates[K].split(" at ")
                    EventDateTemp2=EventDateTemp[1].split()
                    EventDateFinal=EventDateTemp[0]+" "+EventDateTemp2[0]+" "+EventDateTemp2[1]
                    if(EventDateFinal not in Value):
                        print("Error: Excel sheet includes wrong date for event " + EventNames[K])
                        print(Value)
                        print(EventDates[K])
                if(c == 3):
                    if(Value != EventStatuses[K]):
                        print("Error: Excel sheet includes wrong status for event " + EventNames[K])
                        print(Value)
                        print(EventStatuses[K])
                if(c==4):
                    if(Value != str(EventTicketsSold[K])):
                        print("Error: Excel sheet includes wrong tickets sold for event " + EventNames[K])
                        print(Value)
                        print(EventTicketsSold[K])
                if (c==5):
                    if(Value != str(EventTicketsAvailable[K])):
                        print("Error: Excel sheet includes wrong tickets available for event " + EventNames[K])
                        print(Value)
                        print(EventTicketsAvailable[K])
        print()
